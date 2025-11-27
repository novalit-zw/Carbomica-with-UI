from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from pathlib import Path
from typing import Optional, List, Dict, Any
import json
import traceback
import uuid
import subprocess
import os
import shutil
import tempfile

# third-party for Excel handling
import pandas as pd

from storage import (
    create_project_folder,
    save_input_file,
    project_path,
    load_projects_index,
    save_projects_index,
    add_project_to_index,
)
from variables import load_variables, save_variables
from books import generate_books

APP_ORIGINS = [
    "http://localhost:3000",
    "http://127.0.0.1:3000"
]

app = FastAPI(title="Carbonica Engine API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=APP_ORIGINS,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Simple in-memory job store (also persisted to project folder)
_jobs = {}

def _status_file(proj: Path):
    return proj / "outputs" / "status.json"

def _write_status(project_id: str, data: dict):
    proj = project_path(project_id)
    out = proj / "outputs"
    out.mkdir(exist_ok=True, parents=True)
    fp = _status_file(proj)
    fp.write_text(json.dumps(data, indent=2))
    _jobs[project_id] = data

def _read_status(project_id: str):
    proj = project_path(project_id)
    fp = _status_file(proj)
    if fp.exists():
        try:
            return json.loads(fp.read_text())
        except Exception:
            return {"status": "unknown"}
    return _jobs.get(project_id, {"status": "not_started"})

@app.post("/upload")
async def upload_input(file: UploadFile = File(...), project_name: Optional[str] = Form(None)):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files are accepted")
    project_id = create_project_folder(project_name)
    dest = save_input_file(project_id, file)
    # initialize status
    _write_status(project_id, {"status": "uploaded", "input": str(dest)})
    # ensure variables file exists (empty) so frontend can GET immediately
    save_variables(project_id, load_variables(project_id) or {})
    return {"project_id": project_id, "input_path": str(dest)}

@app.post("/upload-temp")
async def upload_temp(file: UploadFile = File(...)):
    """
    Temporary upload endpoint used by the frontend file uploader.
    Does NOT create a project or modify projects.json.
    Saves the uploaded file into a safe uploads/ temp folder and returns a token/filename.
    """
    uploads_dir = Path("uploads")
    uploads_dir.mkdir(exist_ok=True, parents=True)
    # create a unique filename to avoid collision
    suffix = Path(file.filename).suffix
    token = f"{uuid.uuid4()}{suffix}"
    dest = uploads_dir / token
    try:
        with dest.open("wb") as out_f:
            shutil.copyfileobj(file.file, out_f)
    finally:
        await file.close()
    return {"uploaded": True, "token": token, "filename": file.filename, "path": str(dest)}

@app.get("/projects/{project_id}/variables")
def get_variables(project_id: str):
    vars_ = load_variables(project_id)
    return {"variables": vars_}

@app.put("/projects/{project_id}/variables")
def put_variables(project_id: str, payload: dict):
    save_variables(project_id, payload)
    return {"status": "ok", "variables": load_variables(project_id)}

def _call_engine_direct(input_file: Path, out_dir: Path, scenario: Optional[str], options: Optional[dict]):
    """
    Directly call the engine's programmatic entrypoint exposed in run_main.run_project.
    """
    try:
        import run_main
    except Exception as e:
        return False, f"import_run_main_failed: {e}"

    fn = getattr(run_main, "run_project", None)
    if not fn or not callable(fn):
        # fallback tries older names
        for alt in ("run", "run_project_core", "main"):
            fn = getattr(run_main, alt, None)
            if fn and callable(fn):
                break
    if not fn:
        return False, "no_callable_engine_entrypoint"

    try:
        # Prefer signature (input_path, out_dir, scenario, options)
        try:
            res = fn(str(input_file), str(out_dir), scenario, options)
            return True, f"called {fn.__name__} directly, returned: {res}"
        except TypeError:
            # try simpler signatures
            try:
                res = fn(str(input_file), str(out_dir))
                return True, f"called {fn.__name__}(input,out), returned: {res}"
            except TypeError:
                # try no-arg main
                res = fn()
                return True, f"called {fn.__name__}(), returned: {res}"
    except Exception as e:
        return False, f"exception calling engine entrypoint: {e}\n{traceback.format_exc()}"

def _call_engine_subprocess(input_file: Path, out_dir: Path, scenario: Optional[str], options: Optional[dict]):
    cmd = ["python", str(Path(__file__).parent / "run_main.py"), "--input", str(input_file), "--out", str(out_dir)]
    if scenario:
        cmd += ["--scenario", scenario]
    # forward options if provided (spending, budgets)
    if options:
        if "spending" in options and options.get("spending") is not None:
            try:
                cmd += ["--spending", str(float(options.get("spending")))]
            except Exception:
                cmd += ["--spending", str(options.get("spending"))]
        if "budgets" in options and options.get("budgets") is not None:
            b = options.get("budgets")
            if isinstance(b, (list, tuple)):
                bstr = ",".join(str(x) for x in b)
            else:
                bstr = str(b)
            cmd += ["--budgets", bstr]
    proc = subprocess.run(cmd, capture_output=True, text=True)
    return proc.returncode == 0, proc.stdout + proc.stderr

def _run_background(project_id: str, scenario: Optional[str], options: Optional[dict]):
    proj = project_path(project_id)
    # Resolve the actual input file name from variables / index, fallback to any .xls* in project
    inp = None
    try:
        vars_ = load_variables(project_id) or {}
        fn = vars_.get("input_filename")
        if fn:
            candidate = proj / fn
            if candidate.exists():
                inp = candidate
    except Exception:
        inp = None

    if inp is None:
        # try index entry
        try:
            idx = load_projects_index()
            for p in idx.get("projects", []):
                if p.get("project_id") == project_id:
                    fn = p.get("input_filename")
                    if fn:
                        candidate = proj / fn
                        if candidate.exists():
                            inp = candidate
                            break
        except Exception:
            pass

    if inp is None:
        # fallback to first workbook in project
        for f in proj.glob("*.xls*"):
            if f.is_file():
                inp = f
                break

    if inp is None:
        # record failure and exit early
        _write_status(project_id, {"status": "failed", "info": "input workbook not found"})
        return

    out = proj / "outputs"
    out.mkdir(exist_ok=True, parents=True)

    # ensure project env is available for called modules (utils.py will pick this up)
    os.environ["PROJECT_DIR"] = str(proj.resolve())

    _write_status(project_id, {"status": "running", "pid": None, "input": str(inp)})
    # Try direct call first
    ok, info = _call_engine_direct(inp, out, scenario, options)
    if ok:
        _write_status(project_id, {"status": "finished", "info": info})
        return
    # fallback to subprocess (now forwards options)
    ok2, info2 = _call_engine_subprocess(inp, out, scenario, options)
    if ok2:
        _write_status(project_id, {"status": "finished", "info": info2})
    else:
        _write_status(project_id, {"status": "failed", "info": info2})

@app.post("/projects/{project_id}/run")
def run_project(project_id: str, background_tasks: BackgroundTasks, scenario: Optional[str] = None, options: Optional[dict] = None):
    proj = project_path(project_id)
    inp = proj / "input_data.xlsx"
    if not inp.exists():
        raise HTTPException(status_code=404, detail="Input file not found")
    # enqueue background task
    _write_status(project_id, {"status": "queued"})
    background_tasks.add_task(_run_background, project_id, scenario, options)
    return {"status": "queued"}

@app.get("/projects/{project_id}/status")
def project_status(project_id: str):
    return _read_status(project_id)

# helper to list projects — read the persisted index (normalized shape)
@app.get("/projects")
def list_projects() -> Dict[str, Any]:
    idx = load_projects_index()
    return {"projects": idx.get("projects", [])}

# create project (upload + optional start_year + optional project_id)
@app.post("/projects")
async def create_project(
    file: Optional[UploadFile] = File(None),
    project_name: Optional[str] = Form(None),
    start_year: Optional[int] = Form(None),
    project_id: Optional[str] = Form(None),
    upload_token: Optional[str] = Form(None),
):
    """
    Create project. If upload_token is provided, move file from uploads/ into the new project.
    This endpoint is responsible for creating the project index entry (create_project_folder(..., add_index=True)).
    """
    # If a file is provided directly (legacy), accept it; otherwise if upload_token provided, move the temp file.
    if not file and not upload_token:
        raise HTTPException(status_code=400, detail="No file provided or upload_token specified")

    pid = create_project_folder(project_name=project_name, project_id=project_id, add_index=True)
    proj_dir = project_path(pid)

    # handle file move/copy
    if upload_token:
        uploads_dir = Path("uploads")
        src = uploads_dir / upload_token
        if not src.exists():
            raise HTTPException(status_code=400, detail="Upload token not found")
        dest = proj_dir / src.name
        shutil.move(str(src), str(dest))
        saved_path = str(dest)
    else:
        # direct upload stream to project folder
        dest = proj_dir / (file.filename or f"input_{pid}.xlsx")
        try:
            with dest.open("wb") as out_f:
                shutil.copyfileobj(file.file, out_f)
        finally:
            if file:
                await file.close()
        saved_path = str(dest)

    # extract facility_code if available
    try:
        df_fac = pd.read_excel(saved_path, sheet_name="facility", index_col="Code Name")
        facility_code = df_fac.index[0] if len(df_fac.index) > 0 else None
    except Exception:
        facility_code = None

    vars_ = load_variables(pid) or {}
    if project_name:
        vars_["project_name"] = project_name
    if start_year is not None:
        vars_["start_year"] = int(start_year)
    if facility_code:
        vars_["facility_code"] = str(facility_code)
    # record input filename in vars too
    vars_["input_filename"] = Path(saved_path).name
    save_variables(pid, vars_)

    # generate per-project books into projects/{pid}/books
    books_dir = project_path(pid) / "books"
    try:
        gen_start = int(vars_.get("start_year") or start_year or 2024)
        gen_end = int(vars_.get("end_year") or (gen_start + 5))
        generate_books(saved_path, gen_start, gen_end, output_dir=str(books_dir))
    except Exception as e:
        # Don't fail create -- just log and continue; front-end can retry book generation
        idx = load_projects_index()
        # update index with books_path even if generation failed
        for p in idx.get("projects", []):
            if p.get("project_id") == pid:
                p["books_path"] = str(books_dir.resolve())
                break
        save_projects_index(idx)

    # update projects index entry with filename / start_year / name / facility_code / books_path
    idx = load_projects_index()
    updated = False
    for p in idx.get("projects", []):
        if p.get("project_id") == pid:
            p["project_name"] = vars_.get("project_name") or p.get("project_name")
            if start_year is not None:
                p["start_year"] = vars_.get("start_year")
            if facility_code:
                p["facility_code"] = vars_.get("facility_code")
            p["input_filename"] = Path(saved_path).name
            p["books_path"] = str(books_dir.resolve())
            p["has_outputs"] = (project_path(pid) / "outputs").exists()
            updated = True
            break
    if not updated:
        idx.setdefault("projects", []).append({
            "project_id": pid,
            "project_name": vars_.get("project_name"),
            "start_year": vars_.get("start_year"),
            "input_filename": Path(saved_path).name,
            "books_path": str(books_dir.resolve()),
            "created_at": None,
            "has_outputs": (project_path(pid) / "outputs").exists(),
            "facility_code": vars_.get("facility_code"),
        })
    save_projects_index(idx)

    return {
        "project_id": pid,
        "input_path": str(dest),
        "project_name": vars_.get("project_name"),
        "start_year": vars_.get("start_year"),
        "facility_code": vars_.get("facility_code"),
        "books_path": str(books_dir.resolve()),
    }

# get a sheet as JSON (databook/progbook)
@app.get("/projects/{project_id}/sheet")
def get_sheet(project_id: str, sheet: str = "databook", sheet_name: Optional[str] = None):
    proj = project_path(project_id)

    # build candidate workbook list (prefer outputs/{sheet}.xlsx, then recorded/uploaded files, then any .xlsx)
    candidates = []
    out_candidate = proj / "outputs" / f"{sheet}.xlsx"
    if out_candidate.exists():
        candidates.append(out_candidate)

    try:
        idx = load_projects_index()
        for p in idx.get("projects", []):
            if p.get("project_id") == project_id:
                fn = p.get("input_filename")
                if fn:
                    candidates.append(proj / fn)
                break
    except Exception:
        pass

    try:
        vars_ = load_variables(project_id) or {}
        fn = vars_.get("input_filename")
        if fn:
            candidates.append(proj / fn)
    except Exception:
        pass

    if not candidates:
        for f in proj.glob("*.xls*"):
            candidates.append(f)

    # remove duplicates while preserving order
    seen = set()
    candidates = [c for c in candidates if c not in seen and not seen.add(c)]

    if not candidates:
        raise HTTPException(status_code=404, detail="No workbook found for project")

    # If a specific sheet_name requested, find a workbook that contains it.
    requested = sheet_name or None

    chosen_file = None
    chosen_sheet = None

    try:
        if requested:
            req_norm = requested.strip()
            for c in candidates:
                if not c.exists():
                    continue
                try:
                    xlf = pd.ExcelFile(c)
                    # match case-insensitively and trim whitespace
                    match = next((s for s in xlf.sheet_names if s.strip().lower() == req_norm.lower()), None)
                    if match:
                        chosen_file = c
                        chosen_sheet = match
                        break
                except Exception:
                    # skip unreadable files
                    continue
            if not chosen_file:
                # not found in any candidate
                candidate_names = [str(c) for c in candidates if c.exists()]
                raise HTTPException(status_code=404, detail=f"Sheet '{requested}' not found in project workbooks. Candidates: {candidate_names}")
        else:
            # no specific sheet requested: use first existing candidate and use 'sheet' param if it's a sheet name,
            # otherwise read first sheet.
            for c in candidates:
                if c.exists():
                    chosen_file = c
                    break
            if not chosen_file:
                raise HTTPException(status_code=404, detail="No workbook found for project")
            # if 'sheet' looks like a sheet name (non-default 'databook'), try to use it if present
            try:
                xlf = pd.ExcelFile(chosen_file)
                match = next((s for s in xlf.sheet_names if s.strip().lower() == sheet.strip().lower()), None)
                chosen_sheet = match or 0
            except Exception:
                chosen_sheet = 0

        # read the sheet
        df = pd.read_excel(chosen_file, sheet_name=chosen_sheet)
        rows = df.fillna("").to_dict(orient="records")
        return {"sheet": chosen_sheet if isinstance(chosen_sheet, (str, int)) else str(chosen_sheet), "rows": rows, "columns": list(df.columns)}
    except HTTPException:
        raise
    except Exception as e:
        # return a clearer 500 with trace for debugging
        return JSONResponse(status_code=500, content={"error": str(e), "trace": traceback.format_exc()})

# update a sheet (replace whole sheet)
@app.put("/projects/{project_id}/sheet")
def put_sheet(project_id: str, payload: Dict):
    """
    payload: { sheet: "emission data"|'emission targets'|..., sheet_name?: str, rows: [{col:val,..}], columns?: [colnames] }
    This replaces the sheet in outputs/{sheet}.xlsx (or creates it). Also tries to write changes directly into per-project books/{...}
    (databook/progbook) when available; otherwise falls back to editing the uploaded input workbook.
    """
    proj = project_path(project_id)
    data = payload
    sheet = data.get("sheet", "databook")
    sheet_name = data.get("sheet_name", None)
    rows = data.get("rows", [])
    columns = data.get("columns", None)
    if not isinstance(rows, list):
        raise HTTPException(status_code=400, detail="rows must be a list")
    out_dir = proj / "outputs"
    out_dir.mkdir(exist_ok=True, parents=True)
    target = out_dir / f"{sheet}.xlsx"
    try:
        df = pd.DataFrame(rows)
        if columns:
            df = df.reindex(columns=columns)
        # write to outputs/{sheet}.xlsx for quick viewing
        df.to_excel(target, sheet_name=sheet or "Sheet1", index=False, engine="openpyxl")

        # Prefer to write into per-project books if databook/progbook exists
        books_dir = proj / "books"
        written_to_books = False
        if books_dir.exists():
            # try to find an appropriate workbook in books/ that contains the sheet name
            candidate_books = list(books_dir.glob("*.xls*"))
            # heuristic: prefer files that include 'databook'/'progbook' or facility code
            candidate_books.sort(key=lambda p: (('databook' in p.name.lower()) or ('progbook' in p.name.lower())), reverse=True)
            for cb in candidate_books:
                try:
                    xlf = pd.ExcelFile(cb)
                    # match sheet_name preference: sheet_name param -> sheet, else look for sheet in sheet_titles
                    desired = sheet_name or sheet
                    match = next((s for s in xlf.sheet_names if s.strip().lower() == str(desired).strip().lower()), None)
                    # if we found a workbook that already contains the sheet, write into it
                    target_sheet_name = match or (sheet_name or sheet or "Sheet1")
                    # use ExcelWriter replace if supported
                    try:
                        with pd.ExcelWriter(cb, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                            df.to_excel(writer, sheet_name=target_sheet_name, index=False)
                    except ValueError:
                        # fallback remove sheet then append
                        import openpyxl
                        wb = openpyxl.load_workbook(cb)
                        if target_sheet_name in wb.sheetnames:
                            std = wb[target_sheet_name]
                            wb.remove(std)
                            wb.save(cb)
                        with pd.ExcelWriter(cb, engine="openpyxl", mode="a") as writer:
                            df.to_excel(writer, sheet_name=target_sheet_name, index=False)
                    written_to_books = True
                    break
                except Exception:
                    continue

        wrote_to_input = False
        if not written_to_books:
            # Try to write back into the project's uploaded workbook (recorded input_filename or any .xlsx in project)
            input_candidates = []
            try:
                idx = load_projects_index()
                for p in idx.get("projects", []):
                    if p.get("project_id") == project_id:
                        fn = p.get("input_filename")
                        if fn:
                            input_candidates.append(proj / fn)
                        break
            except Exception:
                pass

            try:
                vars_ = load_variables(project_id) or {}
                fn = vars_.get("input_filename")
                if fn:
                    input_candidates.append(proj / fn)
            except Exception:
                pass

            if not input_candidates:
                for f in proj.glob("*.xls*"):
                    input_candidates.append(f)

            input_file = next((c for c in input_candidates if c.exists()), None)

            if input_file:
                # replace or create sheet inside the input workbook
                try:
                    with pd.ExcelWriter(input_file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                        df.to_excel(writer, sheet_name=sheet or "Sheet1", index=False)
                    wrote_to_input = True
                except ValueError:
                    import openpyxl
                    wb = openpyxl.load_workbook(input_file)
                    target_sheet_name = sheet or "Sheet1"
                    if target_sheet_name in wb.sheetnames:
                        std = wb[target_sheet_name]
                        wb.remove(std)
                        wb.save(input_file)
                    with pd.ExcelWriter(input_file, engine="openpyxl", mode="a") as writer:
                        df.to_excel(writer, sheet_name=target_sheet_name, index=False)
                    wrote_to_input = True

        return {"status": "ok", "path": str(target), "wrote_to_books": written_to_books, "wrote_to_input": wrote_to_input}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e), "trace": traceback.format_exc()})


# ---------- Scenarios endpoints ----------
@app.get("/projects/{project_id}/scenarios")
def list_scenarios(project_id: str):
    """
    Returns list of available scenario objects for a project.
    Sources (in order): variables.json 'scenarios' (can be list[str] or list[dict]), outputs/ subfolders (as basic scenarios).
    Each scenario object: { name: str, spending?: number, budgets?: [number,...] }
    """
    proj = project_path(project_id)
    vars_ = load_variables(project_id) or {}
    scenarios_out: List[Dict[str, Any]] = []

    # load from variables.json if present
    vs = vars_.get("scenarios")
    if isinstance(vs, list):
        # accept both list of strings and list of dicts
        for v in vs:
            if isinstance(v, dict) and v.get("name"):
                scenarios_out.append({"name": v.get("name"), **{k: v.get(k) for k in ("spending", "budgets") if k in v}})
            elif isinstance(v, str):
                scenarios_out.append({"name": v})
    # include any directories under outputs as simple scenarios
    out = proj / "outputs"
    if out.exists():
        for p in out.iterdir():
            if p.is_dir():
                # only add if not duplicate
                if not any(s["name"] == p.name for s in scenarios_out):
                    scenarios_out.append({"name": p.name})

    # fallback baseline if empty
    if not scenarios_out:
        scenarios_out = [{"name": "baseline"}]

    return {"scenarios": scenarios_out}


@app.post("/projects/{project_id}/scenarios")
def create_scenario(project_id: str, payload: Dict[str, Any]):
    """
    Create a scenario for the project and persist in variables.json.
    Payload: { name: str, spending?: number, budgets?: [number,...] }
    Returns created scenario object.
    """
    name = payload.get("name")
    if not name:
        raise HTTPException(status_code=400, detail="Scenario name is required")
    spending = payload.get("spending")
    budgets = payload.get("budgets")

    vars_ = load_variables(project_id) or {}
    sc_list = vars_.get("scenarios", [])
    # normalize existing string-list to dict-list
    normalized: List[Dict[str, Any]] = []
    for s in sc_list:
        if isinstance(s, dict) and s.get("name"):
            normalized.append(s)
        elif isinstance(s, str):
            normalized.append({"name": s})
    # prevent duplicates
    if any(s.get("name") == name for s in normalized):
        raise HTTPException(status_code=400, detail="Scenario with that name already exists")
    new = {"name": name}
    if spending is not None:
        try:
            new["spending"] = float(spending)
        except Exception:
            new["spending"] = spending
    if budgets:
        # coerce to list of floats
        try:
            new["budgets"] = [float(x) for x in budgets]
        except Exception:
            new["budgets"] = budgets
    normalized.append(new)
    vars_["scenarios"] = normalized
    save_variables(project_id, vars_)
    return {"scenario": new}

@app.post("/projects/{project_id}/scenarios/run")
def run_scenario(project_id: str, background_tasks: BackgroundTasks, payload: Dict):
    """
    payload: { scenario: 'name', options?: {...} }
    Enqueue background run (uses existing _run_background) and records status.
    """
    scenario = payload.get("scenario") or payload.get("name") or "baseline"
    options = payload.get("options", None)
    # record queued status per scenario
    _write_status(project_id, {"status": "queued", "scenario": scenario})
    background_tasks.add_task(_run_background, project_id, scenario, options)
    return {"status": "queued", "scenario": scenario}


@app.get("/projects/{project_id}/scenarios/{scenario}/table")
def get_scenario_table(project_id: str, scenario: str, sheet: Optional[str] = None):
    """
    Try to find outputs/{scenario}/{sheet}.xlsx => outputs/{sheet}.xlsx => books/* that contains sheet.
    Returns same shape as get_sheet: { sheet, rows, columns }.
    """
    proj = project_path(project_id)
    candidates = []

    # 1) outputs/{scenario}/{sheet}.xlsx
    if (proj / "outputs" / scenario).exists():
        if sheet:
            candidates.append(proj / "outputs" / scenario / f"{sheet}.xlsx")
        else:
            # choose any workbook in scenario folder
            candidates.extend(list((proj / "outputs" / scenario).glob("*.xls*")))

    # 2) outputs/{sheet}.xlsx
    if sheet:
        candidates.append(proj / "outputs" / f"{sheet}.xlsx")

    # 3) project books
    books_dir = proj / "books"
    if books_dir.exists():
        candidates.extend(list(books_dir.glob("*.xls*")))

    candidates = [c for c in candidates if c.exists()]
    if not candidates:
        raise HTTPException(status_code=404, detail="No scenario workbook found")

    # find workbook that contains the sheet (if given)
    chosen_file = None
    chosen_sheet = sheet or None
    try:
        if sheet:
            req_norm = sheet.strip().lower()
            for c in candidates:
                try:
                    xlf = pd.ExcelFile(c)
                    match = next((s for s in xlf.sheet_names if s.strip().lower() == req_norm), None)
                    if match:
                        chosen_file = c
                        chosen_sheet = match
                        break
                except Exception:
                    continue
            if not chosen_file:
                # fallback to first candidate
                chosen_file = candidates[0]
                chosen_sheet = sheet
        else:
            chosen_file = candidates[0]
            chosen_sheet = 0
        df = pd.read_excel(chosen_file, sheet_name=chosen_sheet)
        rows = df.fillna("").to_dict(orient="records")
        return {"sheet": chosen_sheet, "rows": rows, "columns": list(df.columns)}
    except HTTPException:
        raise
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e), "trace": traceback.format_exc()})

@app.get("/projects/{project_id}/books")
def project_books(project_id: str):
    """
    List generated book workbooks for a project.
    Returns: { books_path: str, files: [filename.xlsx, ...] }
    """
    proj = project_path(project_id)
    books_dir = proj / "books"
    try:
        if not books_dir.exists():
            # ensure the directory exists for new projects
            books_dir.mkdir(parents=True, exist_ok=True)
            return {"books_path": str(books_dir.resolve()), "files": []}
        files = sorted([p.name for p in books_dir.glob("*.xls*")])
        return {"books_path": str(books_dir.resolve()), "files": files}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e), "trace": traceback.format_exc()})

@app.get("/projects/{project_id}/books/{filename}/sheets")
def book_sheets(project_id: str, filename: str):
    """
    Return sheet names for a given workbook inside projects/{project_id}/books/{filename}
    """
    proj = project_path(project_id)
    books_dir = proj / "books"
    target = books_dir / filename
    if not target.exists():
        # try decode encoded filename (spaces etc)
        from urllib.parse import unquote
        target2 = books_dir / unquote(filename)
        if target2.exists():
            target = target2
        else:
            raise HTTPException(status_code=404, detail="Workbook not found")
    try:
        xlf = pd.ExcelFile(target)
        return {"filename": target.name, "sheets": xlf.sheet_names}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e), "trace": traceback.format_exc()})

# Serve project graphs list and files
@app.get("/projects/{project_id}/graphs")
def project_graphs(project_id: str):
    proj = project_path(project_id)
    graphs_dir = proj / "graphs"
    try:
        graphs_dir.mkdir(parents=True, exist_ok=True)
        files = sorted([p.name for p in graphs_dir.glob("*") if p.is_file()])
        manifest = graphs_dir / "manifest.json"
        manifest_data = {}
        if manifest.exists():
            try:
                manifest_data = json.loads(manifest.read_text(encoding="utf-8"))
            except Exception:
                manifest_data = {}
        return {"graphs_path": str(graphs_dir.resolve()), "files": files, "manifest": manifest_data}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e), "trace": traceback.format_exc()})

@app.get("/projects/{project_id}/graphs/{filename}")
def serve_graph(project_id: str, filename: str):
    proj = project_path(project_id)
    graphs_dir = proj / "graphs"
    target = graphs_dir / filename
    # try decoding
    if not target.exists():
        from urllib.parse import unquote
        alt = graphs_dir / unquote(filename)
        if alt.exists():
            target = alt
    if not target.exists():
        raise HTTPException(status_code=404, detail="Graph not found")
    return FileResponse(target, media_type="image/png", filename=target.name)